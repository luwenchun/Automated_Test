"""
============================
Author:柠檬班-木森
Time:2020/2/7   20:32
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""

"""
封装的需求：
    1、读取数据
    2、写入数据

封装目的：
    1、提高代码的重用率
    
封装的时候有哪些数据需要参数化？
    1、excel文件名
    2、表单名



"""

import openpyxl


class ReadExcel(object):

    def __init__(self, filename, sheet_name):
        self.filename = filename
        self.sheet_name = sheet_name

    def read_data(self):
        """读取数据"""
        wb = openpyxl.load_workbook(self.filename)
        # 选择表单
        sh = wb[self.sheet_name]
        # 按行获取表单所有格子中的数据，每一行的数据放在一个元组中
        datas = list(sh.rows)
        # 获取第一行的数据，作为字典的键
        title = [i.value for i in datas[0]]
        # 创建一个空列表，用例存放用例数据
        cases = []
        # 遍历除第一行之外的数据
        for i in datas[1:]:
            # 获取该行数据的值
            values = [c.value for c in i]
            # 将该行数据和title（第一行数据）打包转换为字典
            case = dict(zip(title, values))
            # 将转换的字典添加到前面创建的空列表cases中
            cases.append(case)
        return cases

    def write_data(self, row, column, value):
        """写入数据"""
        # 获取工作簿对象
        wb = openpyxl.load_workbook(self.filename)
        # 选择表单
        sh = wb[self.sheet_name]
        # 写入数据
        sh.cell(row=row, column=column, value=value)
        # 保存文件
        wb.save(self.filename)


# excel = ReadExcel("musen1.xlsx", "register")
# case_datas = excel.read_data()
# print("读取出来的数据为：",case_datas)

#
excel1 = ReadExcel("cases.xlsx", "login")
case_datas = excel1.read_data()
# print("读取出来的数据为：",case_datas)


#  注意事项：excel中只有数字和文字（字符串）
case_data = case_datas[0]
# print(case_data["case_id"],type(case_data["case_id"]))
# print(case_data["data"],type(case_data["data"]))
# print(case_data["expected"],type(case_data["expected"]))

# eval：识别字符串中的python表达式
a = case_data["data"]
print(eval(a), type(eval(a)))

# 写入数据
# excel = ReadExcel("musen.xlsx", "name")
# excel.write_data(1, 1, 'python66')  # 位置传参
# excel.write_data(row=2, column=1, value="关键字11")  # 关键字传参
# excel.write_data(row=3, column=1, value="8888")  # 关键字传参
