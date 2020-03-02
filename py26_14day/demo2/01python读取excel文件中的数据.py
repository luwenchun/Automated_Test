"""
============================
Author:柠檬班-木森
Time:2020/2/5   21:20
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""

# 用读写文件的方式直接读取excel文件拿不到想要的数据
# with open("cases.xlsx", "rb", ) as f:
#     content = f.read()
#     print(content)

# 通过openpyxl这个模块来操作
import openpyxl

# 1、数据的读取
# 第一步：读取excel文件，返回一个workbook对象（工作簿）
wb = openpyxl.load_workbook("cases.xlsx")

print(wb)

# 第二步：选取工作簿中的表单
sh = wb["register"]
print(sh)

# 第三步：读取表单中某个格子的内容
data = sh.cell(row=1, column=1).value
print(data)

# 小案例
res = wb["musen"].cell(row=4, column=2).value
print(res)

# 2、数据的写入
# 第一步：读取excel文件，返回一个workbook对象（工作簿）
wb = openpyxl.load_workbook("cases.xlsx")

# 第二步：选取工作簿中的表单
sh = wb["musen"]

# 第三步：写入内容
# sh.cell(row=4, column=3, value="大佬很多")
sh.cell(row=4, column=4).value = "大佬9口罩很多"
# 第四步：保存工作簿为文件
wb.save("cases.xlsx")
