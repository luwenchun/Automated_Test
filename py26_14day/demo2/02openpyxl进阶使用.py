"""
============================
Author:柠檬班-木森
Time:2020/2/5   21:41
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""
import openpyxl

wb = openpyxl.load_workbook("cases.xlsx")

# 表单对象的rows属性
sh = wb["register"]

# 按行获取表单所有格子中的数据，每一行的数据放在一个元组中
datas = list(sh.rows)

"""
[
(<Cell 'register'.A1>, <Cell 'register'.B1>, <Cell 'register'.C1>, <Cell 'register'.D1>, <Cell 'register'.E1>),
(<Cell 'register'.A2>, <Cell 'register'.B2>, <Cell 'register'.C2>, <Cell 'register'.D2>, <Cell 'register'.E2>), 
(<Cell 'register'.A3>, <Cell 'register'.B3>, <Cell 'register'.C3>, <Cell 'register'.D3>, <Cell 'register'.E3>), 
(<Cell 'register'.A4>, <Cell 'register'.B4>, <Cell 'register'.C4>, <Cell 'register'.D4>, <Cell 'register'.E4>),
(<Cell 'register'.A5>, <Cell 'register'.B5>, <Cell 'register'.C5>, <Cell 'register'.D5>, <Cell 'register'.E5>),
(<Cell 'register'.A6>, <Cell 'register'.B6>, <Cell 'register'.C6>, <Cell 'register'.D6>, <Cell 'register'.E6>), 
(<Cell 'register'.A7>, <Cell 'register'.B7>, <Cell 'register'.C7>, <Cell 'register'.D7>, <Cell 'register'.E7>),
(<Cell 'register'.A8>, <Cell 'register'.B8>, <Cell 'register'.C8>, <Cell 'register'.D8>, <Cell 'register'.E8>)
]
"""

# 通过下标获取指定行的数据
# print(datas[2])
li = []
for i in datas[1]:
    # print(i.value)
    li.append(i.value)
print(li)

# 列表推导式
# li = []
# for i in datas[1]:
#     li.append(i.value)
#
# # 列表推导式的语法（python高阶编程的语法）
# li1 = [i.value for i in datas[2]]
# print(li1)


# 2、读取指定列的数据
# 表单的columns属性：按列获取表单所有格子中的数据，每一列的数据放在一个元组中
datas2 = list(sh.columns)
# print(datas2)

# 将指定列的内容读取到一个列表中保存
# li3 = []
# for i in datas2[3]:
#     li3.append(i.value)
# print(li3)

# 通过列表推导式实现
# li4 = [i.value for i in datas2[2]]
# print(li4)


# 3、max_row:获取最大行
print(sh.max_row)


# 4、max_column获取最大列
print(sh.max_column)